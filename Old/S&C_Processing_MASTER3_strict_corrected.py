# -*- coding: utf-8 -*-
"""
Created on Thu Aug 15 14:35:21 2019
Process CMJ, DJ and Pogos .txt files -- CURRENTLY ONLY DJ
@author: BStone
"""
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os.path
import openpyxl
import tkinter as tk
from tkinter import messagebox
import os
import tkinter 
from tkinter import filedialog, Tk
root=Tk()
root.withdraw()
root.wm_attributes('-topmost', 1)
from tkinter.filedialog import askopenfilename, askopenfilenames
from tkinter import simpledialog
import warnings 
warnings.filterwarnings("ignore")
#%% Functions
def locations_of_substring(string, substring):
    """Return a list of locations of a substring."""
        substring_length = len(substring)    
    def recurse(locations_found, start):
    location = string.find(substring, start)
    if location != -1:
    return recurse(locations_found + [location], location+substring_length)
else:
return locations_found
return recurse([], 0)
#%% Select files to loop through
root = tk.Tk()
root.withdraw()
filez = filedialog.askopenfilenames(parent=root, initialdir='D:/Rob/Data/',  title = 'Choose files to process')
rn = len(filez) # Define number of files to loop through
for n in range(rn): # Make Loop
#try:
if rn>1:
filepath = filez[n] # Define filepath based on Nth loop
else:
filepath=filez[0]
with open (filepath, 'r') as f:
data = f.read()
## EXTRACT ADDED DATA    
Fs=1000
Ts = 1/Fs
locs = (locations_of_substring(data, 'Date')) # Find date of trial
date = data[locs[-1]+5:locs[-1]+17]
Dates = ('Jan', 'Feb','Mar', 'Apr','May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')
mth = Dates.index(date[0:3]); mth=mth+1; 
if mth<10:
mth=str(mth); mth='0'+mth
else:
mth=str(mth)# Find month number
day = (date[4:6]) # Find day
yr = (date[-4:len(date)]) # Find year
DATE = day+'/'+mth+'/'+yr
locs = (locations_of_substring(filepath, '/'))
ath = filepath[locs[-1]+1:-1]
locs = (locations_of_substring(ath, ' '))
athlete = ath[0:locs[0]]
if "cmj" in ath:
jump = "CMJ"
elif "dj" in ath:
jump = "DJ"
elif "pogos" in ath:
jump = "POGO"
if ' r ' in ath[locs[0]:len(ath)]:
leg = 'Right'
elif ' R ' in ath[locs[0]:len(ath)]:
leg = 'Right'
elif ' l ' in ath[locs[0]:len(ath)]:
leg = 'Left'
elif ' L ' in ath[locs[0]:len(ath)]:
leg = 'Left'
else:
leg = 'Double'
try:
#%% CMJ PROCESSING
if jump == 'CMJ':  
## EXTRACT FORCE DATA
locs = (locations_of_substring(data, 'N'))
F = data[locs[-1]+2:-1] # Pull force data from complete txt file
F = F.split()
# Define time and force vectors as numpy arrays
t = F[0::7]; t=np.asarray(t) 
F1x = F[1::7]; F1x=np.asarray(F1x); F1x = F1x.astype(float)
F1y = F[2::7]; F1y=np.asarray(F1y); F1y = F1y.astype(float)
F1z = F[3::7]; F1z=np.asarray(F1z); F1z = F1z.astype(float)
F2x = F[4::7]; F2x=np.asarray(F2x); F2x = F2x.astype(float)
F2y = F[5::7]; F2y=np.asarray(F2y); F2y = F2y.astype(float)
F2z = F[6::7]; F2z=np.asarray(F2z); F2z = F2z.astype(float)
Fz = F1z+F2z
BW = np.mean(Fz[0:1200])
BWkg = BW/9.812
res1 = np.sqrt(np.power(Fz-BW,2))
MT = np.max(res1[0:1200])*4
Mv = (Fz > (BW-MT)) & (Fz<(Fz+MT))
SI1 = Fz < BW
SoM1 = np.where(Mv == False)
SoM = SoM1[0][0]
SoMF = Fz[SoM]
if BW>SoMF:
PosNeg = 1
else:
PosNeg = 0
SI2a=np.mean(SI1[0:SoM+1])
for i in range(1, SoM):
SI2a=np.append(SI2a, np.mean(SI1[i:SoM+1]))
for i in range(SoM+1, len(SI1)):
SI2a = np.append(SI2a, np.mean(SI1[SoM:i]))
SI2 = SI2a==1    
offset=float(0)
for n in range(1, len(Fz)):
if (n+360)<len(Fz):
offset = np.append(offset, np.std(Fz[n:n+361], ddof=1))
else:
offset = np.append(offset, np.std(Fz[n:len(Fz)], ddof=1))
offset[-1]=0    
res2 = np.sqrt(np.power(Fz,2))
pkres2=np.amax(res2[0:360])
for n in range(1, len(res2)):
if (n+360)<len(Fz):
pkres2 = np.append(pkres2, np.amax(res2[n:n+361]))
else:
pkres2 = np.append(pkres2, np.amax(res2[n:len(Fz)]))
SoJ = np.where(SI2 == 1)
SoJ = int(SoJ[0][0])
FzJ = Fz[SoJ-1:len(Fz)-999]            
locs = (locations_of_substring(ath,'.'))
FILE=ath[0:locs[0]]
plt.figure(figsize=(15,10))
plt.rc('legend', fontsize=20)
plt.rc('axes', labelsize=20)
plt.rc('xtick', labelsize=20)
plt.rc('ytick', labelsize=20)
plt.plot(FzJ)
plt.show()
plt.xlabel('Data Point')
plt.ylabel('Force (N)')
plt.legend([FILE])
window = tk.Tk()
window.eval('tk::PlaceWindow %s' % window.winfo_toplevel())
window.withdraw()
messagebox.showinfo('Question', 'Press okay when you are happy to continue')
plt.close('all')
#%% Calculate impulse, Velocity, Displacement etc...
imp = (((FzJ[0]+FzJ[1])/2)-BW)*Ts
for n in range(1,(len(FzJ)-1)):
imp = np.append(imp, (((FzJ[n]+FzJ[n+1])/2)-BW)*Ts)
vel = float(0)
vel = np.append(vel,vel+(imp[0]/(BW/9.812)))
for n in range(2, len(imp)+1):
vel = np.append(vel, vel[n-1]+(imp[n-1]/(BW/9.812)))
disp = float(0)
disp = np.append(disp, disp+(vel[0])*Ts)
for n in range(2, len(vel)+1):
disp = np.append(disp, disp[n-1]+(vel[n-1])*Ts)
pwr = FzJ*vel
eccRFD = FzJ[75:-1]
eccRFD1 = eccRFD>BW
eccRFDcheck1 = 'False'
for n in range (1,len(eccRFD1)):
eccRFDcheck1 = np.append(eccRFDcheck1, (eccRFD1 [n-1] == 0) & (eccRFD1 [n] == 1))     
eccRFDcheck2 = 'False'  
for n in range(1, len(eccRFD1)):
if n+10<=len(eccRFD1):
eccRFDcheck2 = np.append(eccRFDcheck2, (eccRFD1[n-1]==0) & (np.sum(eccRFD1[n:n+11]) == 11))
else:
eccRFDcheck2 = np.append(eccRFDcheck2, (eccRFD1[n-1]==0) & (np.sum(eccRFD1[n:n+(len(eccRFD1)-n)])==len(eccRFD1)-n))
#%% Extract parameters
# Peak landing Force
pkLF = np.amax(Fz)
pkLFloc = np.where(Fz == pkLF); pkLFloc = int(pkLFloc[0])
# Minimum offset Force
MinOffsetF = np.amin(offset[SoJ:pkLFloc])
MinOffsetFloc = np.where(offset == MinOffsetF); MinOffsetFloc = int(MinOffsetFloc[0])
# Takeoff threshold
TOThresh = pkres2[MinOffsetFloc]   
#Landing Threshold
LThresh = TOThresh
# Maximum Negative Displacement
MaxNegDisp = np.amin(disp[0:(pkLFloc-(SoJ-1))])
MaxNegDisploc = np.where(disp==MaxNegDisp); MaxNegDisploc = int(MaxNegDisploc[0])
# Location of Takeoff Value
a = Fz[(MaxNegDisploc+(SoJ-1)):-1]
b = np.where(a<=TOThresh)
TOThreshloc = (b[0])+ MaxNegDisploc + (SoJ-1);
TOThreshloc=TOThreshloc[0]
del a, b
# Location of Landing Value
a = Fz[TOThreshloc:pkLFloc]
b = np.where(a<=LThresh)
LThreshloc = (b[0][-1]+TOThreshloc)+1;
del a,b
# Check for 1 sec quiet stance
if SoJ>=1000:
OSQ = 'YES'
else:
OSQ = 'NO'
# Takeoff Velocity and displacement
TOvel = vel[(TOThreshloc-(SoJ-1))]
TOdisp = disp[(TOThreshloc-(SoJ-1))]
# Landing Velocity
Lvel = vel[(LThreshloc-(SoJ-1))]
#Flight time (s)
FT = (LThreshloc-TOThreshloc)/1000
# Find average, peak and relative concentric power (W and W/kg) and contraction time (s)
pkconpwr = np.amax(pwr[1:(LThreshloc-SoM)])
avconpwr = np.mean(pwr[MaxNegDisploc:(TOThreshloc-(SoJ-1)+1)])
conT=(TOThreshloc-SoM)/1000
relpwr = pkconpwr/BWkg
# Jump height
JH = np.power(TOvel,2)/(9.812*2)
#Find maximum eccentric velocity and rise time to this
Maxeccvel = np.amin(vel[1:(TOThreshloc-SoM)]);
Maxeccvelloc = np.where(vel == Maxeccvel);Maxeccvelloc=int(Maxeccvelloc[0])+1
a = FzJ[MaxNegDisploc]
b = FzJ[Maxeccvelloc]
EccRiseRFD = (a-b)/((MaxNegDisploc-Maxeccvelloc)/1000);
del a,b
pkF = np.amax(FzJ[0:(LThreshloc-SoM)]);
pkFloc = np.where(FzJ==pkF); pkFloc=int(pkFloc[0]); pkFloc=pkFloc+SoJ-1
# Max Vel
Maxvel = np.amax(vel[1:(LThreshloc-SoM)]);
#Net Impulse
netimp = np.sum(imp[SoM-(SoJ-1):(TOThreshloc-(SoJ)+2)]);
# Stretch impulse
a = np.asarray(eccRFDcheck2)
a = np.where(eccRFDcheck2 == 'True')
eccRFDpt1loc = a[0][0]+SoJ-1;
eccRFDpt1 = eccRFD[eccRFDpt1loc-(SoJ-1)];
if (eccRFDpt1loc - (SoJ-1))<(MaxNegDisploc):
stretchimp = np.sum(imp[eccRFDpt1loc - (SoJ-1):MaxNegDisploc+1]);
else:
stretchimp = np.sum(imp[MaxNegDisploc:(eccRFDpt1loc - (SoJ-1))]);
del a
# Find positive impulse
endposimp = np.amax(vel[(SoM-(SoJ-1)):(TOThreshloc-(SoJ-1)+1)]);
a =  Fz[MaxNegDisploc+(SoJ-1):-1];
b = np.where(a<=endposimp);
b=b[0][0]-1
endposimploc = b+MaxNegDisploc+SoJ;
if (eccRFDpt1loc-(SoJ-1))<(endposimploc-(SoJ-1)):
posimp = np.sum(imp[eccRFDpt1loc-(SoJ-1):endposimploc-(SoJ-1)+1]);
else:
posimp = np.sum(imp[(endposimploc-(SoJ-1)):(eccRFDpt1loc-(SoJ-1))]);
del a,b
# Average power and eccentric power
a = pwr[SoM-(SoJ-1):MaxNegDisploc+1];
b = np.where(a<0)
aveccpwr = np.mean(a[b])
del a,b
# Power Utilisation
pwrut = avconpwr+aveccpwr
#%% Single plate analysis
# Force from each plate during stance
p1st = np.mean(F1z[0:1200])/BW*100
p2st = np.mean(F2z[0:1200])/BW*100
p1F = F1z[SoJ-1:-999]
p2F = F2z[SoJ-1:-999]
# Min from both plates
p1min = np.amin(p1F[0:MaxNegDisploc])
p2min = np.amin(p2F[0:MaxNegDisploc])
# Max during contraction phase
p1con = np.amax(p1F[0:(TOThreshloc-SoM)])
p2con = np.amax(p2F[0:(TOThreshloc-SoM)])
# Eccentric RFD
p1eccRFD = (p1F[MaxNegDisploc]-p1F[Maxeccvelloc])/((MaxNegDisploc - Maxeccvelloc)/1000);
p2eccRFD = (p2F[MaxNegDisploc]-p2F[Maxeccvelloc])/((MaxNegDisploc - Maxeccvelloc)/1000);
# Mean F from landing to 100ms after
p1Lto100 = np.mean(F1z[(LThreshloc+2):(LThreshloc+103)])
p2Lto100 = np.mean(F2z[(LThreshloc+2):(LThreshloc+103)])
# Peak Landing Force
p1Lmax = np.amax(F1z[(LThreshloc-100):(LThreshloc+502)])
p2Lmax = np.amax(F2z[(LThreshloc-100):(LThreshloc+502)])
#%% Export to Excel sheet 
fname2 = filepath[0:-6]
b1=''
CMJ = np.hstack((DATE, leg, BWkg, relpwr, pkconpwr, JH, EccRiseRFD, pkF, conT, Maxvel, netimp, stretchimp, posimp, MaxNegDisp, aveccpwr, avconpwr, pwrut,b1,p1st, p2st, p1min, p2min, p1con, p2con, p1eccRFD, p2eccRFD, p1Lto100, p2Lto100, p1Lmax, p2Lmax))
CMJ = pd.DataFrame(CMJ); CMJ = CMJ.transpose()
CMJ = CMJ.apply(pd.to_numeric, errors = 'ignore', downcast='float')
# Allow for each leg to be tested
if fname2[-1] == 'j':
fname3 = fname2
else:
fname3 = fname2[0:-2]
path = 'D:/Rob/Data/'
filecheck=path+athlete+'.xlsx'
FC = os.path.exists(filecheck)
if FC==False:
# if file doesnt exist...
header1a = ['Counter Movement Jump Variables','','','','','','','','','','','','','','','','',' ','Single Leg Variables','','','','','','','','','','','']
H1a = pd.DataFrame(header1a); H1a = H1a.transpose()
header1b = ['Single Leg Variables']
H1b = pd.DataFrame(header1b);
header2a = ['Date collected', 'Leg', 'Body Mass (kg)', 'Relative Power (W/kg)','Peak Power (W)', 'Jump Height (m)' ,'aRFD rise total (N/s)', 'Peak Force (N)', 'Contraction Time (s)', 'Peak Velocity (m/s)', 'Net Impulse (N.s)', 'Stretch Impulse (N.s)', 'Positive Impulse (N.S)','Jump Depth (m)', 'Mean Ecc Power', 'Mean Con Power', 'Power Utilisation',' ','Plate 1 Stance', 'Plate 2 Stance', 'Plate 1 Min Force (N)', 'Plate 2 Min Force (N)', 'Plate 1 Max Force (Contraction) (N)', 'Plate 2 Max Force (Contraction) (N)', 'Plate 1 Ecc RFD', 'Plate 2 Ecc RFD', 'Plate 1 landing to 100ms', 'Plate 2 anding to 100ms', 'Plate 1 Landing Peak', 'Plate 2 Landing Peak']
H2a = pd.DataFrame(header2a); H2a = H2a.transpose()
header2b = ['Plate 1 Stance', 'Plate 2 Stance', 'Plate 1 Min Force (N)', 'Plate 2 Min Force (N)', 'Plate 1 Max Force (Contraction) (N)', 'Plate 2 Max Force (Contraction) (N)', 'Plate 1 Ecc RFD' 'Plate 2 Ecc RFD', 'Plate 1 landing to 100ms', 'Plate 2 anding to 100ms', 'Plate 1 Landing Peak', 'Plate 2 Landing Peak']
H2b = pd.DataFrame(header2b); H2b = H2b.transpose()
Heads = pd.concat([H1a, H2a, CMJ], axis=0)
# Convert curves to dataframe to write to excel
writer = pd.ExcelWriter(filecheck, engine='xlsxwriter')
Heads.to_excel(writer, sheet_name='CMJ', index = False, header=False)
writer.save()    
if FC==True:
# IF EXCEL FILE DOES EXIST
xls = pd.ExcelFile(filecheck)
sheets = xls.sheet_names
# CHeck if CMJ tab exists
if sheets.count('CMJ')>0:
check=1
else:
check=0
# If CMJ tab doesnt exist...
if check==0:
wb=openpyxl.load_workbook(filecheck)
try:
wb[date]
except:
wb.create_sheet(date)
book = openpyxl.load_workbook(filecheck)
writer = pd.ExcelWriter(filecheck, engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
CMJ.to_excel(writer, 'CMJ', index=False, header=False)
writer.save()
# If CMJ tab does exist...
else:
currentfile = pd.read_excel(filecheck,sheet_name = 'CMJ', header=None)
c=len(currentfile) # See how many columns are in sheet
CMJ = np.hstack((DATE, leg, BWkg, relpwr, pkconpwr, JH, EccRiseRFD, pkF, conT, Maxvel, netimp, stretchimp, posimp, MaxNegDisp, aveccpwr, avconpwr, pwrut,b1,p1st, p2st, p1min, p2min, p1con, p2con, p1eccRFD, p2eccRFD, p1Lto100, p2Lto100, p1Lmax, p2Lmax))
CMJ = pd.DataFrame(CMJ); CMJ = CMJ.transpose()
CMJ = CMJ.apply(pd.to_numeric, errors = 'ignore', downcast='float')
book = openpyxl.load_workbook(filecheck)
writer = pd.ExcelWriter(filecheck, engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
CMJ.to_excel(writer, 'CMJ', startrow=c, index=False, header=False)
writer.save()
#%% DJ PROCESSING            
elif jump =='DJ':
if len(locs)==4:
BoxH=int(ath[locs[-3]+1:locs[-2]])  
else:
BoxH=int(ath[locs[-2]+1:locs[-1]])  
# EXTRACT FORCE DATA
locs = (locations_of_substring(data, 'N'))
F = data[locs[-1]+2:-1] # Pull force data from complete txt file
F = F.split()
# Define time and force vectors as numpy arrays
t = F[0::7]; t=np.asarray(t) 
F1x = F[1::7]; F1x=np.asarray(F1x); F1x = F1x.astype(float)
F1y = F[2::7]; F1y=np.asarray(F1y); F1y = F1y.astype(float)
F1z = F[3::7]; F1z=np.asarray(F1z); F1z = F1z.astype(float)
F2x = F[4::7]; F2x=np.asarray(F2x); F2x = F2x.astype(float)
F2y = F[5::7]; F2y=np.asarray(F2y); F2y = F2y.astype(float)
F2z = F[6::7]; F2z=np.asarray(F2z); F2z = F2z.astype(float)
P1F = F1z[0:5000]; P2F = F2z[0:5000];
BW = np.mean(P1F[0:1501])
BWkg = BW/9.812
plt.figure(figsize=(15,10))
plt.rc('legend', fontsize=20)
plt.rc('figure', titlesize=20)
plt.rc('axes', labelsize=20)
plt.rc('xtick', labelsize=20)
plt.rc('ytick', labelsize=20)
plt.plot(P1F)
plt.plot(P2F)
plt.show()
plt.xlabel('Data Point')
plt.ylabel('Force (N)')
plt.title(ath[0:-3])
plt.legend(['Plate 1','Plate 2'])
window = Tk()
window.eval('tk::PlaceWindow %s' % window.winfo_toplevel())
window.withdraw()
messagebox.showinfo('Question', 'Press okay when you are happy to continue')
plt.close('all')
res1P1 = np.sqrt(np.power(P1F-BW,2))
resP1 = np.sqrt(np.power(P1F,2)); resP2 = np.sqrt(np.power(P2F,2));
#Calc residuals for P1 and P2 
maxresP1 = np.amax(resP1[0:501])
maxresP2 = np.amax(resP2[0:501])
for j in range(1,len(resP1)):
if j+500<len(resP1):
maxresP1 = np.append(maxresP1, np.amax(resP1[j:j+501]))
maxresP2 = np.append(maxresP2, np.amax(resP2[j:j+501]))
else:
maxresP1 = np.append(maxresP1, np.amax(resP1[j:len(resP1)]))
maxresP2 = np.append(maxresP2, np.amax(resP2[j:len(resP2)]))
#Calc moving averag (P1 and P2)
MAP1 = np.std(P1F[0:501], ddof=1)
MAP2 = np.std(P2F[0:501], ddof=1)
for j in range(1,len(P1F)):
if j+500<len(P1F):
MAP1 = np.append(MAP1, np.std(P1F[j:j+501], ddof=1))
MAP2 = np.append(MAP2, np.std(P2F[j:j+501], ddof=1))
else:
MAP1 = np.append(MAP1, np.std(P1F[j:len(P1F)], ddof=1))
MAP2 = np.append(MAP2, np.std(P2F[j:len(P1F)], ddof=1))
init = np.amax(res1P1[0:1001])*1.75;        
OF1F = np.amin(MAP1[0:4501]) # offset 1
OF1loc = np.where(MAP1==OF1F); OF1loc=OF1loc[0][0]
TO1 = maxresP1[OF1loc]*2
P1init = P1F<TO1
Mv = (P1F > (BW-init)) & (P1F< (BW+init))
P1FBW = P1F<BW
# Find start of movement
SoM = np.where(Mv==False); SoM = SoM[0][0]
SoMF = P1F[SoM]
if BW>SoMF:
PosNeg = 1
else:
PosNeg = 0
if P1FBW[0]==PosNeg:
avP1FBW=1
else:
avP1FBW=0    
for j in range(1,SoM):
if np.mean(P1FBW[j:SoM])==PosNeg:
avP1FBW=np.append(avP1FBW,1)
else:
avP1FBW=np.append(avP1FBW,0)
for j in range(SoM+1, len(P1FBW)):
if np.mean(P1FBW[SoM:j])== PosNeg:
avP1FBW=np.append(avP1FBW,1)
else:
avP1FBW=np.append(avP1FBW,0)
OF2F = np.amin(MAP2[0:4501]) # Offset 2
OF2loc = np.where(MAP2 == OF2F); OF2loc=OF2loc[0][0]
if isinstance(OF2loc, (np.ndarray))=='True':
OF2loc=OF2loc[-1]
LTOP2 = maxresP2[OF2loc]*2
SoJ = np.where(avP1FBW==1)
if SoJ[0][0]>0:
SoJ = SoJ [0][0]-1
else:
SoJ = SoJ [0][1]-1
SoJF = P1F[SoJ]
TO2 = maxresP2[OF2loc]*2
TOnum1 = np.where(P1init==1); TOnum1=TOnum1[0][0]
P2L = ((P2F[0] > TO2) & (P2F[1] > TO2)& (P2F[2] > TO2) & (P2F[3] > TO2) & (P2F[4] > TO2))
for j in range (1,len(P2F)):
if j+5<len(P2F):
P2L = np.append(P2L, ((P2F[j] > TO2) & (P2F[j+1] > TO2)& (P2F[j+2] > TO2) & (P2F[j+3] > TO2) & (P2F[j+4] > TO2)))
Lnum1 = np.where(P2L == 1); Lnum1 = Lnum1[0][0]
# Make Fz array that has both force plates in divided at right point
Fz = P1F[0:TOnum1]
Fz = np.append(Fz,0)
Fz = np.append(Fz, P2F[(TOnum1+1):len(P2F)])
# Calc vertical F, acc, vel and disp during jump
FzJ = Fz[SoJ:len(Fz)]
acc = (FzJ-BW)/(BW/9.812)
vel = 0
vel=np.append(vel,(vel+((acc[0]*(1/Fs)))))
for j in range(2,(len(acc)+1)):
vel=np.append(vel,(vel[j-1]+((acc[j-1]*(1/Fs)))))
imp=0
for j in range(1,len(FzJ)):
imp=np.append(imp, (FzJ[j]-BW)*(1/Fs)+(0.5*((FzJ[j]-FzJ[j-1])*(1/Fs))))
disp=0
disp=np.append(disp, (disp+(vel[0]*(1/Fs))))
for j in range(2,len(vel)+1):
disp=np.append(disp, (disp[j-1]+(vel[j-1]*(1/Fs))))
vel=vel[0:-1]    
pwr = FzJ * vel
WD = FzJ*disp[0:-2]
FBW = FzJ/BW
MAF = np.std(FzJ[0:11], ddof=1)
for j in range(1,len(FzJ)):
if j+11<len(FzJ):
MAF = np.append(MAF,np.std(FzJ[j:j+11], ddof=1) )
else:
MAF=np.append(MAF, np.std(FzJ[j:-1], ddof=1))
TO2 = np.where(P2L[Lnum1:-1]==0); TO2 = TO2[0][0]+Lnum1
Lnum2 = np.where(P2L[TO2:-1]==1); Lnum2 = Lnum2[0][0]+TO2+1
# Displacement @ key points
maxnegdisp = np.amin(disp[(Lnum1-SoJ):(TO2-SoJ)])
maxnegdisploc=np.where(disp==maxnegdisp) + SoJ
maxnegdisploc=maxnegdisploc[0][0]+1
dispTO1 = disp[(TOnum1-SoJ)-1]
dispL1 = disp[(Lnum1-SoJ)-1]
dispCT = maxnegdisp-dispL1
dispTO2 = disp[(TO2-SoJ)-1]
dispL2 = disp[(Lnum2-SoJ)-1]
#Vel @ certain points
VTO1 = vel[(TOnum1-SoJ)-1]
VL1 = vel [(Lnum1-SoJ)-1]
VTO2 = vel[(TO2-SoJ)-1]            
# Times of different phases
CT = (TO2-Lnum1)/1000    # Contact time
EP = (maxnegdisploc-Lnum1)/1000  # Eccentric phase
CP = (TO2 - maxnegdisploc)/1000  # Concentric phase
FT = (Lnum2 - TO2)/1000   # Flight time
# Jump height based on velocity
JHv = (np.power(VTO2,2)/(9.81*2))
JHt = (9.812*(np.power(FT,2)))/8
# Calc forces at key points
maxnegdispF = FzJ[(maxnegdisploc-SoJ)-1]
EccF = np.amax(FzJ[Lnum1-SoJ:maxnegdisploc-SoJ])
EccFloc = np.where(FzJ==EccF); EccFloc=EccFloc+SoJ; EccFloc = EccFloc[0][0]+1
ConF = np.amax(FzJ[maxnegdisploc-SoJ:TO2-SoJ])
ConFloc = np.where(FzJ==ConF); ConFloc=ConFloc[0][0]+SoJ+1
AvEccF=np.mean(FzJ[((Lnum1-SoJ)-1):((maxnegdisploc-SoJ))])
AvConF = np.mean(FzJ[((maxnegdisploc-SoJ)-1):TO2-SoJ])
EccconF=np.amax(EccF)
EccconF=np.append(EccconF, np.amax(ConF))
PkF = np.amax(EccconF)
PkFloc = np.where(FzJ==PkF); PkFloc=PkFloc[0][0]+SoJ+1
EccFBW = EccF/BW
ConFBW = ConF/BW
# Time to key points
EccFt = (EccFloc-Lnum1)/1000
ConFt = ((ConFloc-Lnum1)/1000)-EccFt
# Phase ratios
EccConP=CP/EP
FTCT = FT/CT
# Peak Power
PkEccpwr = np.amin(pwr[Lnum1-SoJ:maxnegdisploc-SoJ])
PkConpwr = np.amax(pwr[maxnegdisploc-SoJ:TO2-SoJ])
PkConpwrloc = np.where(pwr==PkConpwr); PkConpwrloc=PkConpwrloc[0][0]+SoJ+1
Vk = np.abs(maxnegdispF/(maxnegdisp-dispL1))   # Vertical stiffness
RSt = JHt/CT     # Reactive strength based on FT
RSv = JHv/CT     # Reactive strength based on velocity
# Actual drop height
DH = disp[(TOnum1-SoJ)-1]
actDH = (DH*100)+BoxH
effDH = (np.power(VL1,2)/(2*9.81))   # Effective drop height
Pkpwrt = (PkConpwrloc-Lnum1+1)/1000;
avRPD = PkConpwr/Pkpwrt
# Calc Impulse
Conimp = np.sum(imp[((maxnegdisploc-SoJ)-1):TO2-SoJ])
Eccimp = np.sum(imp[((Lnum1-SoJ)-1):maxnegdisploc-SoJ])
Totimp = Eccimp+Conimp
imprat = Eccimp/Conimp
ms50 = Lnum1+50
ms100 = ms50+50
imp50ms= np.sum(imp[((Lnum1-SoJ)-1):ms50-SoJ])
imp100ms= np.sum(imp[((Lnum1-SoJ)-1):ms100-SoJ])
#Average Eccentric power
avEccpwr = np.mean(pwr[((Lnum1-SoJ)-1):maxnegdisploc-SoJ])
avConpwr = np.mean(pwr[((maxnegdisploc-SoJ)-1):TO2-SoJ])
relpwr = PkConpwr/BWkg
pwrut = avEccpwr+avConpwr
#%% Export to Excel sheet
# Export to Excel sheet
# Debugging checks before data preparation
print(f"Checking critical data sizes before processing:")
print(f"Length of maxresP1: {len(maxresP1) if 'maxresP1' in locals() else 'Not defined'}")
print(f"Length of maxresP2: {len(maxresP2) if 'maxresP2' in locals() else 'Not defined'}")
print(f"Length of pwr: {len(pwr) if 'pwr' in locals() else 'Not defined'}")
print(f"Length of WD: {len(WD) if 'WD' in locals() else 'Not defined'}")
# Check if critical arrays are empty before proceeding
if 'maxresP1' in locals() and len(maxresP1) == 0:
raise ValueError("maxresP1 is empty. Data extraction failed at some point.")
if 'maxresP2' in locals() and len(maxresP2) == 0:
raise ValueError("maxresP2 is empty. Data extraction failed at some point.")
if 'pwr' in locals() and len(pwr) == 0:
raise ValueError("pwr is empty. Data extraction failed at some point.")
if 'WD' in locals() and len(WD) == 0:
raise ValueError("WD is empty. Data extraction failed at some point.")
# Prepare data for export
results = {
'Date': [DATE],
'Athlete': [athlete],
'Leg': [leg],
'Box Height': [BoxH],
'Body Weight (kg)': [BWkg],
'Contact Time (s)': [CT],
'Eccentric Time (s)': [EP],
'Concentric Time (s)': [CP],
'Flight Time (s)': [FT],
'Jump Height (m)': [JHv],
'Disp during Contact (m)': [dispCT],
'Actual Drop Height (m)': [actDH],
'Reactive Strength Index': [RSv],
'Velocity @ TO (m/s)': [VTO2],
'Vertical Stiffness (N/m)': [Vk],
'Relative Power (W/kg)': [relpwr],
'Peak Power (W)': [PkConpwr],
'Mean Eccentric Power (W)': [avEccpwr],
'Mean Concentric Power (W)': [avConpwr],
'Power Utilisation (W)': [pwrut],
'Peak Eccentric Force (N)': [EccF],
'Peak Concentric Force (N)': [ConF],
'Eccentric Force/BW': [EccFBW],
'Concentric Force/BW': [ConFBW],
'Total Impulse (N.s)': [Totimp],
'Eccentric Impulse (N.s)': [Eccimp],
'Concentric Impulse (N.s)': [Conimp],
'Impulse @ 50 ms (N.s)': [imp50ms],
'Eccentric:Concentric Impulse Ratio': [imprat],
'Max Residual P1': [np.amax(maxresP1)],
'Max Residual P2': [np.amax(maxresP2)],
'Power': [np.amax(pwr)],
'Work Done': [np.sum(WD)],
}
df_results = pd.DataFrame(results)
# Export to Excel
output_path = 'D:/Rob/Results/'
output_file = os.path.join(output_path, f'{athlete}_DJ_Results.xlsx')
if not os.path.exists(output_file):
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
df_results.to_excel(writer, index=False, sheet_name='DJ Results')
else:
with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
df_results.to_excel(writer, index=False, sheet_name='DJ Results', startrow=writer.sheets['DJ Results'].max_row, header=False)
except Exception as e:
print(f'An error occurred: {e}')
